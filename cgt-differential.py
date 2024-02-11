
# tax rate on dividends is highest marginal combined national/state rate taking account of imputation systems, tax credits and tax allowances. Source: https://stats.oecd.org/index.aspx?DataSetCode=TABLE_II4, column K
# labour income is combined national/state including employee social security (not employer). Source https://stats.oecd.org/Index.aspx?DataSetCode=TABLE_I7 "all in rate" column F
# CGT rate is highest marginal combined national/state rate on shares. Source: https://taxsummaries.pwc.com/quick-charts/capital-gains-tax-cgt-rates and Tax Policy Associates

# these are rates for local residents. Some countries (e.g. France) in principle tax foreigners on local gains (although tax treaties mean in practice it rarely applies). Other countries like the UK/US only tax local residents on gains in stock/shares

import plotly.graph_objs as go
from openpyxl import load_workbook
from PIL import Image

wb = load_workbook(filename='IT_CGT_top_marginal_rates.xlsx')
sheet = wb['CGT_vs_IT']


for income_tax_type in ["dividend", "employment"]:

    # Initialize lists to hold your data
    countries = []
    relevant_it_rates = []
    cgt_rates = []

    # The data starts at row 2 to skip the headers and is in columns A, B, C
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        countries.append(row[0])
        relevant_it_rates.append(row[1 if income_tax_type == "employment" else 2])
        cgt_rates.append(row[3])


    # Sort the list of tuples by the CGT rate in descending order (highest first) 
    country_rate_tuples = list(zip(countries, relevant_it_rates, cgt_rates))
    sorted_country_rate_tuples = sorted(country_rate_tuples, key=lambda x: x[2], reverse=True)
    countries, relevant_it_rates, cgt_rates = zip(*sorted_country_rate_tuples)

    # do logo
    LOGO_JPG_FILE = Image.open("logo_full_white_on_blue.jpg")
    logo_layout = [dict(
            source=LOGO_JPG_FILE,
            xref="paper", yref="paper",
            x=1, y=1.03,
            sizex=0.1, sizey=0.1,
            xanchor="right", yanchor="bottom"
        )]

    fig = go.Figure()

    # Add arrows for each country
    for i, (country, relevant_it_rate, cgt_rate) in enumerate(zip(countries, relevant_it_rates, cgt_rates)):
        
        
        # add IT/dividend rate
        fig.add_shape(
            type="line",
            x0=i,
            y0=relevant_it_rate,
            x1=i,
            y1=cgt_rate,
            line=dict(
                color="red" if country == "United Kingdom" else "blue",
                width=3
            ),
            xref="x",
            yref="y"
        )
        
        # Add arrowhead
        fig.add_trace(go.Scatter(
            x=[i],
            y=[cgt_rate],
            mode='markers',
            marker=dict(
                color="red" if country == "United Kingdom" else "blue",
                size=10,
                symbol='arrow-bar-up' if cgt_rate > relevant_it_rate else 'arrow-bar-down'
            ),
            showlegend=False
        ))
        

    # Set the layout of the figure
    fig.update_layout(
        images=logo_layout,
        title=f'Differential between rates of {"income tax on dividends" if income_tax_type == "dividend" else "income tax and NI/SS on employment income"} and capital gains tax, by country (arrowhead is CGT)',
        xaxis=dict(
            tickmode='array',
            tickvals=list(range(len(countries))),
            ticktext=countries
        ),
        yaxis=dict(
            title='Tax Rate (%)',
            range=[0, max(relevant_it_rates) + 0.05], 
            tickmode='array',
            tickvals=[i/10 for i in range(11)],  
            ticktext=[f'{i*10}%' for i in range(11)], 
            tickformat=',.0%'
        ),
        height=800,
        width=1400
    )

    # Show the figure
    fig.show()